import openpyxl

file_out = open('Fab_cable_correction', 'w')

from  openpyxl  import  load_workbook
wb = load_workbook('calc.xlsx')
sheet = wb['Данные для импорта']

popravki = [[0] * 38 for i in range(12)]



for i in range(12):
    for j in range(38):
        row = sheet.cell(row = (i + 1), column = (j + 1)).value
        popravki[i][j] = row

for i in range(12):
    for j in range(38):
        file_out.write(str(popravki[i][j]) + ' ')
    file_out.write('\n')

file_out.close()